import pandas
import openpyxl
import re
import pandas
from pandas import DataFrame


class HyperlinkToFilename:

    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)
        self.worksheet = self.workbook.active

    @property
    def attachmentColumns(self):
        '''
        附件列序号
        '''
        attachmentColumnsList = []

        for i in range(1, self.worksheet.max_column + 1):
            if '附件' in self.worksheet.cell(1, i).value:
                attachmentColumnsList.append(i)

        return attachmentColumnsList

    def attachmentHyperlinkToFilename(self):

        regexPattern = re.compile(
            r'https\:\/\/wj.qq.com\/api\/files\/download\?survey\_id\=4165188\&question\_id\=(?:.*)\&file\_name\=(.*)\&download\=1', re.I)

        for column in self.attachmentColumns:

            # 首行附件列名称
            currentColumn = self.worksheet.max_column + 1
            self.worksheet.cell(row=1, column=currentColumn).value = self.worksheet.cell(
                row=1, column=column).value

            # 迭代处理附件名
            for row in range(1, self.worksheet.max_row + 1):
                value = self.worksheet.cell(row=row, column=column).value

                if value is not None:
                    match = regexPattern.search(value)
                    if match is not None:
                        # hyperlink = match.group(0)  # 超链接
                        attachName = match.group(1)  # 附件文件名

                        self.worksheet.cell(
                            row=row, column=currentColumn).value = attachName

    def worksheetToDataFrame(self):
        self.dataFrame = DataFrame(self.worksheet.values)
        print(self.dataFrame.columns.values)
        print(self.dataFrame['开始答题时间'])

        # 处理日期时间
        self.dataFrame['开始答题时间'] = pandas.to_datetime(self.dataFrame['开始答题时间'])
        self.dataFrame['结束答题时间'] = pandas.to_datetime(self.dataFrame['结束答题时间'])

        # 将 NA 单元格填充为 ''
        self.dataFrame = self.dataFrame.fillna('')



if __name__ == '__main__':
    # 执行入口

    h = HyperlinkToFilename(r'C:\Users\yuanl\OneDrive\桌面\4165188_202012241706123309.xlsx')

    h.attachmentHyperlinkToFilename()
    # h.worksheetToDataFrame()
    # print(h.dataFrame)
    
    h.workbook.save(r'C:\Users\yuanl\OneDrive\桌面\1.xlsx')
